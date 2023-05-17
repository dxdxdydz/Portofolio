import requests
import json
import openpyxl
import datetime
import time
url = 'https://gql.tokopedia.com/'


def ColIdxToXlName(idx):
    if idx < 1:
        raise ValueError("Index is too small")
    result = ""
    while True:
        if idx > 26:
            idx, r = divmod(idx - 1, 26)
            result = chr(r + ord('A')) + result
        else:
            return chr(idx + ord('A') - 1) + result


def updater(url_name,shopDomain,sid):
    print('Scraping......')
    page = 0
    flag = True
    price = dict()
    sold = dict()
    while flag == True:
        page = page + 1
        paramData = [{"operationName": "ShopProducts",
                      "variables": {"sid": str(sid), "page": page, "perPage": 80, "keyword": "", "etalaseId": "etalase",


      #REDACTED#


            try:
                sold[int(i['product_id'])] = i['label_groups'][0]['title']
            except:
                pass
        if k[0]['data']['GetShopProduct']['links']['next'] == "":
            flag = False
    print('Done Scraping pages!')
    print('Saving.....')
    wb = openpyxl.load_workbook("price_report_"+shopDomain+".xlsx")
    ws = wb.active
    ws.cell(column=19 + ws['D3'].value, row=8, value=datetime.datetime.now())



    #REDACTED#


            ws.cell(column=19 + ws['D3'].value, row=9 + i, value=None)
            ws.cell(column=19, row=row + i, value=None)
        try:
            change.append(((ws['S' + str(row + i)].value - ws['R' + str(row + i)].value)/ws['R' + str(row + i)].value,i))
        except:
            change.append((0, i))

    ws.cell(column=2, row=8, value=datetime.datetime.now())
    ws.cell(column=2, row=8).number_format = 'd-mmm-yy'
    i=0
    change.sort(reverse = True)
    for percent, num in change:
        ws.cell(column=1, row=9 + i, value=ws['Q'+str(num+9)].value)
        ws.cell(column=2, row=9 + i, value=percent)
        ws['B'+str(9+i)].number_format = '0.00%'
        i+=1


    ws['D3'] = ws['D3'].value + 1
    wb.save("price_report_"+shopDomain+".xlsx")


def get_url_name_shopDomain(shopName):
    n = 0
    url_name = []
    wb = openpyxl.load_workbook("price_report_"+shopName+".xlsx")
    ws = wb.active
    while ws['q' + str(9 + n)].value is not None:
        url_name.append(ws['q' + str(9 + n)].value)
        n += 1
    shopDomain = ws['A2'].value
    return url_name, shopDomain

stores=[

    ("bardistore",5275068),
    ("ardushopid",1974041),
     ]

for store in stores:

    start=time.time()
    url_name, shopDomain = get_url_name_shopDomain(store[0])
    updater(url_name, shopDomain, store[1])
    stop=time.time()
    print(shopDomain+" saved in "+str(stop-start)+"s!")